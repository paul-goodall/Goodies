import os
import sys
import cv2
import math
import glob
import json
import time
import copy
import random
import imutils
import statistics
import numpy as np
import pandas as pd
import xlsxwriter
from openpyxl import load_workbook
from astropy.io import fits

from pathlib import Path

import pickle
import bz2
import _pickle as cPickle

import io

pd.set_option('mode.chained_assignment', None)
#
# ==============================================================================
#
script_path = Path(__file__).resolve()
script_dir, script_name = os.path.split(script_path)
#
# ==============================================================================
#
def mkdir(dirpath):
    if not os.path.exists(dirpath):
        os.mkdir(dirpath)
#
# ==============================================================================
#
def glob_files(fpat):
    ff = glob.glob(fpat)
    ff.sort()
    return ff

#
# ==============================================================================
# RAW TXT

def rTXT(outfile):
    with open(outfile, 'r') as f_txt:
        content = f_txt.read()
        f_txt.close()
    return content

def wTXT(data, outfile, append=False):
    fmode = 'w'
    if append:
        fmode = 'a'
    with open(outfile, fmode) as f_txt:
        print(data, file=f_txt)

#
# ==============================================================================
# JSON

def dict2json(dicty, my_indent=4):
    # convert dictionary to JSON string
    json_string = json.dumps(dicty, indent=my_indent)
    return json_string

def rJSON(fname):
    json_string = rTXT(fname)
    json_data   = json.loads(json_string)
    return json_data

def wJSON(json_data, outfile):
    if type(json_data) == dict:
        json_data = dict2json(json_data)
    wTXT(json_data, outfile)

#
# ==============================================================================
# PKLS
# read a pickle:
def rPKL(filename,pickle_string=None):

    if filename == 'from_string':
        return pickle.loads(pickle_string.encode())

    s3 = False
    if 's3://' in filename:
        s3 = True

    if s3:
        fs = s3fs.S3FileSystem(anon=False)
        data = pickle.load(fs.open(filename, 'rb'))
    else:
        if filename[-4:] == 'pbz2':
            data = bz2.BZ2File(filename, 'rb')
            data = cPickle.load(data)
        else:
            dbfile = open(filename, 'rb')
            data = pickle.load(dbfile)
            dbfile.close()
    return data

# write a pickle:
def wPKL(data, filename):

    compress=False
    if filename[-4:] == 'pbz2':
        compress=True

    if filename == 'to_string':
        return pickle.dumps(data, 0).decode()

    s3 = False
    if 's3://' in filename:
        s3 = True

    if s3:
        fs = s3fs.S3FileSystem(anon=False)
        pickle.dump(data, fs.open(filename, 'wb'))
    else:
        if compress:
            with bz2.BZ2File(filename, 'w') as f:
                cPickle.dump(data, f)
        else:
            if os.path.exists(filename):
                os.remove(filename)
            dbfile = open(filename, 'ab')
            pickle.dump(data, dbfile)
            dbfile.close()


# ==============================================================================

def rXL(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    sheet_names = wb.sheetnames
    obj = {}
    for sn in sheet_names:
        obj[sn] = pd.read_excel(filepath, sheet_name=sn)
    return obj

# ==============================================================================

def wXL(obj, filepath):
    print(filepath)
    with pd.ExcelWriter(filepath, engine='xlsxwriter') as writer:
        for sn in list(obj.keys()):
            obj[sn].to_excel(writer, sheet_name=sn, index=False)


# ==============================================================================
# FITS

def wFITS(data, filename, hdr=None, overwrite=True, output_verify='exception'):
    if hdr is None:
        fits.writeto(filename, data, overwrite=overwrite, output_verify=output_verify)
    elif type(hdr) is dict:
        hdr2 = fits.Header()
        for key in hdr:
            hdr2[key] = hdr[key]
        fits.writeto(filename, data, hdr2, overwrite=overwrite, output_verify=output_verify)
    else:
        # Assume here it's a FITS header object:
        print('Got astropy.io hdr')
        fits.writeto(filename, data, hdr, overwrite=overwrite, output_verify=output_verify)

def rFITS(filename):
    data, hdr = fits.getdata(filename, 0, header=True)
    return(data, hdr)

#
# ==============================================================================
#

#
# ==============================================================================
#
def determine_image_depth(im):
    max_vals = np.array([1,255,65535])
    img_max = im.max()
    delta = abs(max_vals - img_max)
    im_threshold = max_vals[np.where(delta == min(delta))][0]
    im_threshold = im_threshold.astype(np.float32) * 1.0
    return (im_threshold)
#
# ==============================================================================
#

def rIMG(filename, switch_rgb=True, normalise=True,flip_y = True):
    im_suffix = os.path.splitext(filename)[-1]
    x = cv2.imread(filename, -1)
    if flip_y:
        x = np.flipud(x)
    x = x[:,:,0:3]
    if switch_rgb:
        x = cv2.cvtColor(x, cv2.COLOR_RGB2BGR)
    if normalise:
        col_max = determine_image_depth(x)
        x = x.astype(np.float32)/col_max
    return x
#
# ==============================================================================
#
def wIMG(x, filename, switch_rgb=True, depth=16,flip_y = True):
    col_max = determine_image_depth(x)
    #print(col_max)
    im_suffix = os.path.splitext(filename)[-1]

    if im_suffix == '.png':
        if depth not in [8,16]:
            depth = 16
    if im_suffix == '.tif':
        if depth not in [8,16,32]:
            depth = 16
    if im_suffix in ['.jpg','.jpeg']:
        depth = 8

    if flip_y:
        x = np.flipud(x)

    x = x.astype(np.float32)/col_max
    if switch_rgb:
        x = cv2.cvtColor(x, cv2.COLOR_RGB2BGR)

    if depth == 8:
        x = (x*255).astype(np.uint8)
    if depth == 16:
        x = (x*65535).astype(np.uint16)
    if depth == 32:
        x = (x*1.0).astype(np.float32)
    cv2.imwrite(filename, x)
#
# ==============================================================================
#

def df2html(df0, html_file, my_page_title='My Smart Table'):

    df = df0.copy()
    search_col = [ ' '.join([str(x) for x in row.tolist()]) for ind, row in df.iterrows() ]
    # make this the first column in the data frame:
    df.insert(0, 'search_col', search_col, allow_duplicates=True)

    my_table_header = '<tr class="header">\n'
    for cn in list(df.columns):
        my_table_header += f'<th>{cn}</th>\n'
    my_table_header += '</tr>'

    my_table_rows = ''
    for ind, row in df.iterrows():
        my_table_rows += '<tr>\n'
        for cn in list(df.columns):
            my_table_rows += f'<td>{row[cn]}</td>\n'
        my_table_rows += '</tr>\n'

    my_table_contents = my_table_header + '\n' + my_table_rows

    smart_table_html = '''
    <!DOCTYPE html>
    <html>
    <head>
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
    * {
      box-sizing: border-box;
    }

    #myInput {
      background-image: url('/css/searchicon.png');
      background-position: 10px 10px;
      background-repeat: no-repeat;
      width: 100%;
      font-size: 16px;
      padding: 12px 20px 12px 40px;
      border: 1px solid #ddd;
      margin-bottom: 12px;
    }

    #myTable {
      border-collapse: collapse;
      width: 100%;
      border: 1px solid #ddd;
      font-size: 18px;
    }

    #myTable th, #myTable td {
      text-align: left;
      padding: 12px;
    }

    #myTable tr {
      border-bottom: 1px solid #ddd;
    }


    #myTable tr.header, #myTable tr:hover {
      background-color: #f1f1f1;
    }

    /* HIDE the first column in the table, which is the search column */
    #myTable tr th:first-child {
        display: none;
    }
    #myTable tr td:first-child {
        display: none;
    }

    </style>
    </head>
    <body>

    <h2>my_page_title</h2>

    <input type="text" id="myInput" onkeyup="myFunction()" placeholder="Search for names.." title="Type in a name">

    <table id="myTable">
      my_table_contents
    </table>

    <script>
    function myFunction() {
      var input, filter, table, tr, td, i, txtValue;
      input = document.getElementById("myInput");
      filter = input.value.toUpperCase();
      table = document.getElementById("myTable");
      tr = table.getElementsByTagName("tr");
      for (i = 0; i < tr.length; i++) {
        td = tr[i].getElementsByTagName("td")[0];
        if (td) {
          txtValue = td.textContent || td.innerText;
          if (txtValue.toUpperCase().indexOf(filter) > -1) {
            tr[i].style.display = "";
          } else {
            tr[i].style.display = "none";
          }
        }
      }
    }
    </script>

    </body>
    </html>
    '''

    smart_table_html = smart_table_html.replace('my_page_title', my_page_title)
    smart_table_html = smart_table_html.replace('my_table_contents', my_table_contents)

    wTXT(smart_table_html, html_file)

#
# ==============================================================================
#

#
# ==============================================================================
#
